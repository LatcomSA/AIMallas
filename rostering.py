# -*- coding: utf-8 -*-
"""
Created on Wed Aug 26 19:36:02 2020

@author: JassonM0lina
"""
import datetime
import openpyxl
from math import nan
import numpy as np
import pandas as pd
import openpyxl 

def rostrng_info(rostrng_total,enc_dec,interval,lunch_hour_total, train_hour_total,block_break,sched_agent,agent_active):
    
    entry = enc_dec[1]
    departure = enc_dec[2]
    break1 = enc_dec[3]
    break2 = enc_dec[4]
    training = enc_dec[5]
    lunch = enc_dec[6]
    
    
    block_lunch = int((lunch_hour_total[0]*60+lunch_hour_total[1])/interval)
    block_training = int((train_hour_total[0]*60+train_hour_total[1])/interval)
    
    days = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday']
    total_auxs = ['preturn','turn','training','break1','break2','lunch']
    drtn = ['init','stop']
    
    for d in rostrng_total:  
        dim = rostrng_total.get(d)
        for x in range(len(dim)):
            entry_count = 0
            break1_count = 0
            break2_count = 0
            training_count = 0
            lunch_count = 0
            caux_sched =  {}
            for y in range(dim[x].shape[0]):
                if dim[x][y] == entry:
                    h1,m1 = divmod(y*interval,60)
                    if y == 0:           
                        h,m = divmod(((60*24)/interval-1)*interval,60)
                    else:   
                        h,m = divmod((y-1)*interval,60)
                    caux_sched[0]=datetime.time(h,m)
                    caux_sched[1]=nan
                    caux_sched[2]=datetime.time(h1,m1)
                    entry_count += 1
                elif dim[x][y] == departure:
                    h,m = divmod((y+1)*interval,60)
                    caux_sched[3] = datetime.time(h,m)
                elif dim[x][y] == training[0]:   
                    h,m = divmod(y*interval,60)
                    h1,m1 = divmod((y+block_training)*interval,60)
                    caux_sched[4] = datetime.time(h,m)
                    caux_sched[5] = datetime.time(h1,m1)
                    training_count += 1
                elif dim[x][y] == break1:
                    h,m = divmod(y*interval,60)
                    h1,m1 = divmod((y+block_break)*interval,60)
                    caux_sched[6] = datetime.time(h,m)
                    caux_sched[7] = datetime.time(h1,m1)
                    break1_count += 1
                elif dim[x][y] == break2:
                    h,m = divmod(y*interval,60)
                    h1,m1 = divmod((y+block_break)*interval,60)
                    caux_sched[8] =datetime.time(h,m)
                    caux_sched[9] = datetime.time(h1,m1)
                    break2_count += 1
                elif dim[x][y] == lunch[0]:   
                    h,m = divmod(y*interval,60)
                    h1,m1 = divmod((y+block_lunch)*interval,60)
                    caux_sched[10]=datetime.time(h,m)
                    caux_sched[11]=datetime.time(h1,m1)
                    lunch_count += 1
            if entry_count == 0:
                for c in range((len(total_auxs)*len(drtn))):
                    caux_sched[c] = nan  
            if break1_count == 0:
                caux_sched[6] = nan 
                caux_sched[7] = nan 
            if break2_count == 0:
                caux_sched[8] = nan 
                caux_sched[9] = nan
            if training_count == 0:
                caux_sched[4] = nan 
                caux_sched[5] = nan
            if lunch_count == 0:
                caux_sched[10] = nan 
                caux_sched[11] = nan             
    
            general_count =0
            for t in total_auxs:
                sched_agent.get(x).graph.add_edge("{}_{}".format(days[d],sched_agent.get(x).name),"{}_{}_{}".format(t,days[d],sched_agent.get(x).name))
                for m in drtn:
                    sched_agent.get(x).graph.add_edge("{}_{}_{}".format(t,days[d],sched_agent.get(x).name),"{}_{}_{}_{}".format(t,days[d],sched_agent.get(x).name,m), time = caux_sched[general_count])
                    general_count += 1
                if t == 'preturn':
                    sched_agent.get(x).graph.remove_node("{}_{}_{}_stop".format(t,days[d],sched_agent.get(x).name))            
    
    
    respta_info = np.zeros([agent_active.shape[0]*len(days),((len(total_auxs)*len(drtn))-1)+2],dtype='O')
    agent_count = 0
    for z in sched_agent:
        for d in days:
           respta_info[agent_count][0] =  sched_agent.get(z).name
           respta_info[agent_count][1] =  d           
           count_auxs = 2
           for t in total_auxs:
               for m in drtn:
                   try:                       
                       respta_info[agent_count][count_auxs] = sched_agent.get(z).graph["{}_{}_{}".format(t,d,sched_agent.get(z).name)]["{}_{}_{}_{}".format(t,d,sched_agent.get(z).name,m)]['time']
                       count_auxs +=1
                   except:
                       continue
           agent_count += 1
    
    wb = openpyxl.Workbook()
    sheet = wb.active    
    
    sheet.append(('name','day','preturno', 'ingreso','salida','out capa','in capa','sal break1','ent break2','sal break2','sal break2','ent lunch', 'sal lunch'))          
    for respta in respta_info:
        sheet.append(tuple(respta))
        
    wb.save('./Rostering/visual_info.xlsx')

def rostrng_act(sched_agent,agent_active):
    
    days = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday']
    total_auxs = ['preturn','turn','training','break1','break2','lunch']
    drtn = ['init','stop']
    
    df = pd.read_excel('./Rostering/visual_info.xlsx', sheet_name='Sheet').to_numpy()
    
    agent_name = []
    new_sched = []
    for nvt in range(agent_active.shape[0]):
        if agent_active[:,1][nvt] == 'ESPECIAL':
           agent_name.append(nvt) 
           pos = np.where(df == agent_active[:,0][nvt])
           for y in pos[0]:
               for x in df[y][2:]:
                   new_sched.append(x)

    count_info = 0
    for nvt in agent_name:
        for d in days:
            for t in total_auxs:
                for m in drtn:
                    try:
                        sched_agent.get(nvt).graph["{}_{}_{}".format(t,d,sched_agent.get(nvt).name)]["{}_{}_{}_{}".format(t,d,sched_agent.get(nvt).name,m)]['time'] = new_sched[count_info]
                        count_info += 1
                    except:
                        continue
    return sched_agent

def rostrng_doc(sched_agent,date):
    
    days = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday']
    total_auxs = ['preturn','turn','training','break1','break2','lunch']
    drtn = ['init','stop']
    
    date_init = datetime.date(date[2],date[1],date[0])
    
 
    up_massive = []
    train_massive = [] 
    for agent in sched_agent.values():
        for t in total_auxs:
            for d in range(len(days)):
                try:
                    if t == 'training':
                       train_init = agent.graph["training_{}_{}".format(days[d],agent.name)]["training_{}_{}_init".format(days[d],agent.name)]['time'] 
                       if train_init is nan:
                            continue
                       else:
                            train_row = np.zeros([1,6],dtype='O')
                            train_row[0,0] = agent.name
                            new_date = date_init + datetime.timedelta(days=d)               
                            train_row[0,2]= new_date.strftime('%d/%m/%Y')
                            train_stop = agent.graph["training_{}_{}".format(days[d],agent.name)]["training_{}_{}_stop".format(days[d],agent.name)]['time']                 
                            train_init = train_init.strftime('%H:%M')
                            train_stop = train_stop.strftime('%H:%M')
                            train_row[0,3] = "{} - {}".format(train_init,train_stop)
                            train_row[0,4] = "Por asignar"
                            train_row[0,5] = "Refuerzo semanal"
                            train_massive.append(train_row)
                    else:                           
                        time = agent.graph["{}_{}_{}".format(t,days[d],agent.name)]["{}_{}_{}_init".format(t,days[d],agent.name)]['time']
                        if time is nan:
                            continue
                        else:
                            row_massive = np.zeros([1,7],dtype='O')
                            row_massive[0,0]= agent.name
                            new_date = date_init + datetime.timedelta(days=d)               
                            row_massive[0,2]= new_date.strftime('%d/%m/%Y')
                            row_massive[0,3]= new_date.strftime('%d/%m/%Y')
                            turn_init = agent.graph["turn_{}_{}".format(days[d],agent.name)]["turn_{}_{}_init".format(days[d],agent.name)]['time'] 
                            turn_stop = agent.graph["turn_{}_{}".format(days[d],agent.name)]["turn_{}_{}_stop".format(days[d],agent.name)]['time']                 
                            turn_init = turn_init.strftime('%H:%M')
                            turn_stop = turn_stop.strftime('%H:%M')
                            row_massive[0,4] = "{} - {}".format(turn_init,turn_stop)
                            row_massive[0,5] = t
                            row_massive[0,6] = time
                            up_massive.append(row_massive) 
                except:
                    continue
    
    subida_masiva = openpyxl.Workbook()
    sheet_subida = subida_masiva.active
    sheet_subida.append(('Nombre','Documento','Fecha Horario', 'Fecha Otro Horario','Horario','Tipo Otro Horario','Horario Otro Horario'))
    for massive in up_massive:
        sheet_subida.append(tuple(massive[0]))
    subida_masiva.save('./Rostering/subida_masiva.xlsx')
    
    archivo_cap = openpyxl.Workbook()
    sheet_archivo = archivo_cap.active
    sheet_archivo.append(('Nombre','Documento','Fecha Horario', 'Horario Capacitacion','Ubicacion Capacitacion','Seccion Capacitacion'))
    for massive2 in train_massive:
        sheet_archivo.append(tuple(massive2[0]))
    archivo_cap.save('./Rostering/archivo_cap.xlsx')
    

    
                    # if t == 'training':
                #     try:
                #         row_massive_cap[0,0] = agent.name
                #         new_date = date_init + datetime.timedelta(days=d)               
                #         row_massive_cap[0,2]= new_date.strftime('%d/%m/%Y')
                #         train_init = agent.graph["training_{}_{}".format(days[d],agent.name)]["training_{}_{}_init".format(days[d],agent.name)]['time'] 
                #         train_stop = agent.graph["training_{}_{}".format(days[d],agent.name)]["training_{}_{}_stop".format(days[d],agent.name)]['time']                 
                #         train_init = train_init.strftime('%H:%M')
                #         train_stop = train_stop.strftime('%H:%M')
                #         row_massive_cap[0,4] = "Por asignar"
                #         row_massive_cap[0,5] = "Refuerzo semanal"                        
                #     except:
                #         continue
                #else:
    
    # wb2 = openpyxl.Workbook()
    # sheet2 = wb2.active
    # sheet2.append(('Nombre','Documento','Fecha Horario', 'Horario Capacitacion','Ubicacion Capacitacion','Seccion Capacitacion'))
    # for massive2 in train_massive:
    #     sheet2.append(tuple(massive2[0]))
    # wb.save('./Rostering/cap.xlsx')
    
  #sched_agent.get(z).graph["{}_{}_{}".format(t,d,sched_agent.get(z).name)]["{}_{}_{}_{}".format(t,d,sched_agent.get(z).name,m)]['time']                
    
    # respta_info = np.zeros([agent_active.shape[0]*len(days),((len(total_auxs)*len(drtn))-1)+2],dtype='O')
    # agent_count = 0
    # for z in sched_agent:
    #     for d in days:
    #        respta_info[agent_count][0] =  sched_agent.get(z).name
    #        respta_info[agent_count][1] =  d           
    #        count_auxs = 2
    #        for t in total_auxs:
    #            for m in drtn:
    #                try:                       
    #                    respta_info[agent_count][count_auxs] = sched_agent.get(z).graph["{}_{}_{}".format(t,d,sched_agent.get(z).name)]["{}_{}_{}_{}".format(t,d,sched_agent.get(z).name,m)]['time']
    #                    count_auxs +=1
    #                except:
    #                    continue
    #        agent_count += 1
    
    # now8= datetime.date(date[2],date[1],date[0])
    # now8= now8.strftime('%d/%m/%Y')
    # now5 = now5.strftime('%H:%M')
    # rty = "{}-{}".format(now5,now4)
    
    # import datetime   
    # import pandas as pd
    # import openpyxl         
    # wb = openpyxl.Workbook()
    # sheet = wb.active      
    # now2 = datetime.datetime(2020, 8, 27)
    # now3 = datetime.time(13, 30)
    # now4 = datetime.time(19, 0)
    # now5 ="{}:{}-{}:{}".format(now3.hour,now3.minute,now4.hour,now4.minute)
    # sheet['A1'].value = now5       
    # wb.save('./Rostering/subida.xlsx')       
    
    

